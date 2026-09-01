from pathlib import Path
from datetime import date
import sys

import pandas as pd

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


sys.path.insert(
    0,
    str(Path(__file__).resolve().parents[1] / "src")
)

from services.chart_service import save_chart
from services.news_service import get_news


TRACKING_FILE = Path(
    "data/tracking/j_type_tracking.csv"
)

RESULT_DIR = Path("results")


SHEET_NAME = "\u004a\u578b"

HEADER_CODE = "\u30b3\u30fc\u30c9"
HEADER_NAME = "\u9298\u67c4\u540d"
HEADER_CLOSE = "\u7d42\u5024"
HEADER_CHANGE1 = "\u524d\u65e5\u6bd4(%)"
HEADER_VOL5 = "\u51fa\u6765\u9ad8/\u76f4\u524d5\u65e5\u5e73\u5747"
HEADER_NEWS = "\u30cb\u30e5\u30fc\u30b9"
HEADER_NEWS_DATE = "\u30cb\u30e5\u30fc\u30b9\u65e5\u6642"
HEADER_CHART = "\u65e5\u8db3\u30c1\u30e3\u30fc\u30c8"


def normalize_code(value):
    if pd.isna(value):
        return ""

    value = str(value).strip()

    if value.endswith(".0"):
        value = value[:-2]

    return value


def get_latest_news(code, name):
    try:
        news_list = get_news(
            code=code,
            name=str(name),
            hours=72,
            limit=5,
        )
    except Exception as exc:
        print(
            f"News ERROR {code}: {exc}"
        )
        return {
            "title": "",
            "published": "",
            "link": "",
        }

    if not news_list:
        return {
            "title": "",
            "published": "",
            "link": "",
        }

    valid = [
        item
        for item in news_list
        if isinstance(item, dict)
    ]

    if not valid:
        return {
            "title": "",
            "published": "",
            "link": "",
        }

    valid.sort(
        key=lambda item: str(
            item.get("published", "")
        ),
        reverse=True,
    )

    return valid[0]


def main():
    today = date.today().isoformat()

    if not TRACKING_FILE.exists():
        print(
            "Tracking file not found:",
            TRACKING_FILE,
        )
        return

    df = pd.read_csv(
        TRACKING_FILE,
        encoding="utf-8-sig",
        low_memory=False,
    )

    required = [
        "JVersion",
        "DetectionDate",
        "Code",
        "Name",
        "BasePrice",
        "Change1",
        "VolumeRatio",
        "DetectionVolumeVsPre5",
    ]

    missing = [
        col
        for col in required
        if col not in df.columns
    ]

    if missing:
        print(
            "Missing columns:",
            missing,
        )
        return

    today_df = df[
        df["DetectionDate"].astype(str)
        == today
    ].copy()

    today_df = today_df[
        today_df["JVersion"].astype(str)
        == "J1"
    ].copy()

    if today_df.empty:
        print(
            "No J1 candidates today:",
            today,
        )
        return

    today_df["Code"] = (
        today_df["Code"]
        .apply(normalize_code)
    )

    today_df = today_df.sort_values(
        "DetectionVolumeVsPre5",
        ascending=False,
    ).reset_index(drop=True)

    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    output_file = (
        RESULT_DIR
        / f"{today}_j_type.xlsx"
    )

    print("=" * 72)
    print("J1 CHART EXCEL")
    print("=" * 72)
    print("Date          :", today)
    print("J1 candidates :", len(today_df))
    print()

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = SHEET_NAME

    headers = [
        HEADER_CODE,
        HEADER_NAME,
        HEADER_CLOSE,
        HEADER_CHANGE1,
        "VolumeRatio",
        HEADER_VOL5,
        HEADER_NEWS,
        HEADER_NEWS_DATE,
        HEADER_CHART,
    ]

    for col_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = sheet.cell(
            row=1,
            column=col_index,
            value=header,
        )

        cell.font = Font(
            bold=True,
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:I{len(today_df) + 1}"
    )

    chart_col = 9

    chart_success = 0
    chart_error = 0

    for index, row in today_df.iterrows():
        excel_row = index + 2

        code = normalize_code(
            row["Code"]
        )

        values = [
            code,
            row["Name"],
            row["BasePrice"],
            row["Change1"],
            row["VolumeRatio"],
            row["DetectionVolumeVsPre5"],
        ]

        for col_index, value in enumerate(
            values,
            start=1,
        ):
            sheet.cell(
                row=excel_row,
                column=col_index,
                value=value,
            )

        news = get_latest_news(
            code,
            row["Name"],
        )

        news_title = str(
            news.get("title", "") or ""
        )

        news_published = str(
            news.get("published", "") or ""
        )

        news_link = str(
            news.get("link", "") or ""
        )

        news_cell = sheet.cell(
            row=excel_row,
            column=7,
            value=news_title,
        )

        sheet.cell(
            row=excel_row,
            column=8,
            value=news_published,
        )

        if news_link:
            news_cell.hyperlink = news_link
            news_cell.style = "Hyperlink"

        news_cell.alignment = Alignment(
            vertical="center",
            horizontal="left",
            wrap_text=True,
        )

        sheet.cell(
            row=excel_row,
            column=8,
        ).alignment = Alignment(
            vertical="center",
        )

        sheet.cell(
            row=excel_row,
            column=1,
        ).number_format = "@"

        for col_index in [
            3,
            4,
            5,
            6,
        ]:
            sheet.cell(
                row=excel_row,
                column=col_index,
            ).number_format = "0.00"

        for col_index in range(
            1,
            chart_col,
        ):
            sheet.cell(
                row=excel_row,
                column=col_index,
            ).alignment = Alignment(
                vertical="center",
            )

        sheet.row_dimensions[
            excel_row
        ].height = 125

        try:
            chart_path = save_chart(
                code
            )

            if chart_path:
                image = Image(
                    str(chart_path)
                )

                image.width = 350
                image.height = 160

                image.anchor = sheet.cell(
                    row=excel_row,
                    column=chart_col,
                ).coordinate

                sheet.add_image(
                    image
                )

                chart_success += 1

            else:
                chart_error += 1

        except Exception as exc:
            chart_error += 1

            print(
                f"Chart ERROR {code}: {exc}"
            )

        print(
            f"Chart {index + 1}/{len(today_df)} : {code}"
        )

    widths = {
        1: 9,
        2: 15,
        3: 9,
        4: 9,
        5: 10,
        6: 13,
        7: 28,
        8: 16,
        9: 40,
    }

    for col_index, width in widths.items():
        sheet.column_dimensions[
            get_column_letter(col_index)
        ].width = width

    sheet.row_dimensions[1].height = 24

    # ======================================
    # Final cell alignment
    # ======================================

    for row_no in range(
        2,
        len(today_df) + 2,
    ):
        for col_no in range(
            1,
            10,
        ):
            cell = sheet.cell(
                row=row_no,
                column=col_no,
            )

            if col_no == 7:
                cell.alignment = Alignment(
                    vertical="top",
                    horizontal="left",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    vertical="top",
                )


    workbook.save(
        output_file
    )

    print()
    print("=" * 72)
    print("J1 Excel complete")
    print("=" * 72)
    print("Charts        :", chart_success)
    print("Chart errors  :", chart_error)
    print("Saved         :", output_file)


if __name__ == "__main__":
    main()
