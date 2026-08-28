import pandas as pd
from datetime import datetime
from pathlib import Path

from openpyxl import (
    Workbook,
    load_workbook,
)
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

from services.chart_service import save_chart

from services.news_service import (
    get_top20_news,
    analyze_news_reason,
)

from services.tracking_service import (
    load_tracking,
    update_tracking_results,
    record_initial_move,
)

from services.p5_tracking_service import (
    load_p5_tracking,
    record_p5_candidates,
    update_p5_tracking,
)

from services.yahoo_credit_service import (
    load_latest_credit_data,
    download_credit_batch,
)

# ==========================================================
# 共通設定
# ==========================================================

RESULT_DIR = Path("results")
CHART_DIR = RESULT_DIR / "charts"


# ==========================================================
# 安全な値取得
# ==========================================================

def get_value(row, column, default=""):
    """
    DataFrameの列が存在しない場合でも落ちない安全な値取得。
    """

    if row is None:
        return default

    if column not in row.index:
        return default

    value = row[column]

    if pd.isna(value):
        return default

    return value


# ==========================================================
# Excel列幅
# ==========================================================

def set_column_widths(sheet, widths):
    """
    Excel列幅を設定。
    """

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


# ==========================================================
# ヘッダー書式
# ==========================================================

def format_header(sheet):
    """
    1行目をヘッダーとして整形。
    """

    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


# ==========================================================
# 本文書式
# ==========================================================

def format_body(sheet):
    """
    本文セルを整形。
    """

    for row in sheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )


# ==========================================================
# 初動スコア色付け
# ==========================================================

def apply_score_color(
    sheet,
    header_name,
):
    """
    初動スコアを低→高で色付けする。
    """

    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    if header_name not in headers:
        return

    score_col = headers[
        header_name
    ]

    scores = []

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        value = sheet.cell(
            row_number,
            score_col
        ).value

        try:

            score = float(value)

            if pd.notna(score):
                scores.append(score)

        except Exception:

            continue

    if not scores:
        return

    min_score = min(scores)
    max_score = max(scores)

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        cell = sheet.cell(
            row_number,
            score_col
        )

        try:

            score = float(
                cell.value
            )

        except Exception:

            continue

        if max_score == min_score:

            ratio = 1.0

        else:

            ratio = (
                score - min_score
            ) / (
                max_score - min_score
            )

        ratio = max(
            0.0,
            min(
                1.0,
                ratio
            )
        )

        red = 255

        green = int(
            255 * (1 - ratio)
        )

        color = (
            f"FF{green:02X}00"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            start_color=color,
            end_color=color,
        )

        cell.font = Font(
            bold=True
        )

def apply_tracking_change_color(
    sheet
):
    """
    初動追跡シートの騰落率を色付けする。

    5%以上10%未満 : 青太字
    10%以上       : 赤太字
    """

    headers = {}

    for cell in sheet[1]:
        headers[cell.value] = cell.column

    change_columns = [
        column_number
        for header_name, column_number in headers.items()
        if (
            isinstance(
                header_name,
                str
            )
            and header_name.endswith(
                "騰落率"
            )
        )
    ]

    if not change_columns:
        return

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        for column_number in change_columns:

            cell = sheet.cell(
                row=row_number,
                column=column_number
            )

            try:

                value = float(
                    cell.value
                )

            except Exception:

                continue

            if pd.isna(value):
                continue

            if value >= 10:

                cell.font = Font(
                    color="FF0000",
                    bold=True
                )

            elif value >= 5:

                cell.font = Font(
                    color="0000FF",
                    bold=True
                )

# ==========================================================
# 初動スコアTOP20作成
# ==========================================================

def make_initial_move_top20(df):
    """
    初動スコアだけでTOP20を作成。

    強気度はランキングに使用しない。
    """

    if df is None:
        return pd.DataFrame()

    if df.empty:
        return df.copy()

    if "初動スコア" not in df.columns:

        print(
            "初動スコア列が存在しません。"
        )

        return pd.DataFrame()

    score = pd.to_numeric(
        df["初動スコア"],
        errors="coerce"
    )

    result = (
        df.loc[
            score.notna()
        ]
        .copy()
    )

    if result.empty:
        return result

    result["_score_numeric"] = (
        pd.to_numeric(
            result["初動スコア"],
            errors="coerce"
        )
    )

    result = (
        result
        .sort_values(
            "_score_numeric",
            ascending=False,
            kind="mergesort",
        )
        .head(20)
        .copy()
    )

    result.drop(
        columns=[
            "_score_numeric"
        ],
        inplace=True,
        errors="ignore",
    )

    return result

# ==========================================================
# P5早期候補抽出
# ==========================================================

def make_p5_candidates(df):
    """
    全銘柄から正式P5条件を満たす銘柄を抽出する。

    P5条件そのものは _build_buy_avoidance_map() で判定し、
    ここでは判定ロジックを重複して持たない。
    """

    if df is None:
        return pd.DataFrame()

    if df.empty:
        return pd.DataFrame()

    if "コード" not in df.columns:
        return pd.DataFrame()

    p5_map = _build_buy_avoidance_map(df)

    p5_mask = []

    for _, row in df.iterrows():

        code = str(
            row.get(
                "コード",
                "",
            )
        ).strip()

        if code.endswith(".0"):
            code = code[:-2]

        info = p5_map.get(
            code,
            {},
        )

        p5_mask.append(
            bool(
                info.get(
                    "p5_condition",
                    False,
                )
            )
        )

    p5_candidates = df.loc[
        p5_mask
    ].copy()

    if p5_candidates.empty:
        return p5_candidates

    # 初動スコア優先、
    # 同点では5日騰落率が高い順
    sort_columns = []
    ascending = []

    if "初動スコア" in p5_candidates.columns:
        sort_columns.append(
            "初動スコア"
        )
        ascending.append(False)

    if "5日騰落率" in p5_candidates.columns:
        sort_columns.append(
            "5日騰落率"
        )
        ascending.append(False)

    if sort_columns:
        p5_candidates = (
            p5_candidates
            .sort_values(
                sort_columns,
                ascending=ascending,
            )
        )

    return p5_candidates.reset_index(
        drop=True
    )

# ==========================================================
# ニュース理由から安全に値を取得
# ==========================================================

def get_reason_value(
    reason,
    keys,
    default="",
):
    """
    analyze_news_reason() の戻り値について
    キー名の違いを吸収する。
    """

    if not isinstance(
        reason,
        dict
    ):
        return default

    for key in keys:

        value = reason.get(
            key,
            ""
        )

        if value is None:
            continue

        if isinstance(
            value,
            float
        ) and pd.isna(value):

            continue

        text = str(value).strip()

        if text:
            return value

    return default


# ==========================================================
# ニュースから最初のURLを取得
# ==========================================================

def get_first_news_url(
    news_list,
):
    """
    ニュース一覧から最初のURLを取得。
    """

    if not isinstance(
        news_list,
        list
    ):
        return ""

    for news in news_list:

        if not isinstance(
            news,
            dict
        ):
            continue

        url = news.get(
            "url",
            news.get(
                "link",
                ""
            )
        )

        if url:
            return str(url)

    return ""


# ==========================================================
# ニュースからタイトルを取得
# ==========================================================

def get_first_news_title(
    news_list,
):
    """
    ニュース一覧から最初のタイトル。
    """

    if not isinstance(
        news_list,
        list
    ):
        return ""

    for news in news_list:

        if not isinstance(
            news,
            dict
        ):
            continue

        title = news.get(
            "title",
            ""
        )

        if title:
            return str(title)

    return ""


# ==========================================================
# ニュースから概要を取得
# ==========================================================

def get_first_news_summary(
    news_list,
):
    """
    ニュース一覧から最初の概要。
    """

    if not isinstance(
        news_list,
        list
    ):
        return ""

    for news in news_list:

        if not isinstance(
            news,
            dict
        ):
            continue

        summary = news.get(
            "summary",
            news.get(
                "description",
                ""
            )
        )

        if summary:
            return str(summary)

    return ""


# ==========================================================
# ニュースから日時を取得
# ==========================================================

def get_first_news_date(
    news_list,
):
    """
    ニュース一覧から最初の日時。
    """

    if not isinstance(
        news_list,
        list
    ):
        return ""

    for news in news_list:

        if not isinstance(
            news,
            dict
        ):
            continue

        published = news.get(
            "published",
            news.get(
                "date",
                ""
            )
        )

        if published:
            return str(
                published
            )

    return ""


# ==========================================================
# チャート作成
# ==========================================================

def create_charts(
    initial_move_top20,
):
    """
    初動スコアTOP20の日足チャートを作成。
    """

    chart_files = {}

    if initial_move_top20.empty:
        return chart_files

    for _, row in initial_move_top20.iterrows():

        code = str(
            get_value(
                row,
                "コード",
                ""
            )
        ).replace(".0", "")

        if not code:
            continue

        try:

            chart = save_chart(
                code
            )

        except Exception as e:

            print(
                f"チャート作成ERROR {code} : {e}"
            )

            continue

        if chart:

            chart_files[
                code
            ] = Path(
                chart
            )

    print(
        f"日足チャート作成完了 ({len(chart_files)}銘柄)"
    )

    return chart_files


# ==========================================================
# DataFrame → Excel
# ==========================================================

def dataframe_to_sheet(
    workbook,
    sheet_name,
    dataframe,
):
    """
    DataFrameをExcelシートに出力。
    """

    if sheet_name in workbook.sheetnames:

        del workbook[
            sheet_name
        ]

    sheet = workbook.create_sheet(
        sheet_name
    )

    if dataframe is None:
        dataframe = pd.DataFrame()

    if dataframe.empty:
        return sheet

    # ヘッダー
    for col_index, column in enumerate(
        dataframe.columns,
        start=1
    ):

        sheet.cell(
            1,
            col_index
        ).value = column

    # データ
    for row_index, (_, row) in enumerate(
        dataframe.iterrows(),
        start=2
    ):

        for col_index, column in enumerate(
            dataframe.columns,
            start=1
        ):

            value = row[column]

            if pd.isna(value):
                value = ""

            sheet.cell(
                row_index,
                col_index
            ).value = value

    format_header(
        sheet
    )

    format_body(
        sheet
    )

    sheet.freeze_panes = "A2"

    sheet.auto_filter.ref = (
        sheet.dimensions
    )

    return sheet


# ==========================================================
# 初動スコアTOP20 Excel 信用情報添付
# ==========================================================


# ==========================================================
# Buy avoidance alert
#
# This alert is completely independent from Initial Score.
#
# A : high-zone stall
# C : one-day spike
# D : extreme overheating
# F : deceleration after previous-day surge
# H2: low score + weak 20-day volume
# ==========================================================

def _alert_float(value):
    try:
        if value is None or pd.isna(value):
            return float("nan")
        return float(value)
    except Exception:
        return float("nan")


def _load_previous_chg1_map():
    """
    Load CHG1 from the most recent stock_result.csv
    before today.

    Used only for F alert.
    """

    today = pd.Timestamp(
        datetime.now().date()
    )

    previous_file = None
    previous_date = None

    for p in RESULT_DIR.glob(
        "*_stock_result.csv"
    ):
        try:
            d = pd.Timestamp(
                p.name[:10]
            )
        except Exception:
            continue

        if d >= today:
            continue

        if (
            previous_date is None
            or d > previous_date
        ):
            previous_date = d
            previous_file = p

    if previous_file is None:
        return {}

    try:
        prev = pd.read_csv(
            previous_file,
            encoding="utf-8-sig",
            dtype=str,
        )
    except Exception as e:
        print(
            f"previous result load ERROR: {e}"
        )
        return {}

    if len(prev.columns) <= 4:
        return {}

    code_col = prev.columns[0]
    chg1_col = prev.columns[4]

    chg1_numeric = pd.to_numeric(
        prev[chg1_col],
        errors="coerce",
    )

    result = {}

    for i, row in prev.iterrows():

        code = str(
            row[code_col]
        ).replace(
            ".0",
            ""
        ).strip()

        if not code:
            continue

        result[code] = (
            chg1_numeric.iloc[i]
        )

    return result


def _build_buy_avoidance_map(
    initial_move_top20,
):
    """
    Return:
        {
            code: {
                "avoid": "回避" or "－",
                "comment": "...",
            }
        }

    Initial Score itself is NOT changed.
    """

    result = {}

    if initial_move_top20 is None:
        return result

    if initial_move_top20.empty:
        return result

    prev_chg1_map = (
        _load_previous_chg1_map()
    )

    for _, row in (
        initial_move_top20.iterrows()
    ):

        code = str(
            get_value(
                row,
                "\u30b3\u30fc\u30c9",
                "",
            )
        ).replace(
            ".0",
            ""
        ).strip()

        chg1 = _alert_float(
            get_value(
                row,
                "\u524d\u65e5\u6bd4",
                float("nan"),
            )
        )

        chg5 = _alert_float(
            get_value(
                row,
                "5\u65e5\u9a30\u843d\u7387",
                float("nan"),
            )
        )

        chg20 = _alert_float(
            get_value(
                row,
                "20\u65e5\u9a30\u843d\u7387",
                float("nan"),
            )
        )

        rsi = _alert_float(
            get_value(
                row,
                "RSI",
                float("nan"),
            )
        )

        score = _alert_float(
            get_value(
                row,
                "\u521d\u52d5\u30b9\u30b3\u30a2",
                float("nan"),
            )
        )

        vol = _alert_float(
            get_value(
                row,
                "VolumeRatio",
                float("nan"),
            )
        )

        vol20 = _alert_float(
            get_value(
                row,
                "VolumeRatio20",
                float("nan"),
            )
        )

        ma25dev = _alert_float(
            get_value(
                row,
                "MA25Deviation",
                float("nan"),
            )
        )

        prev_chg1 = (
            prev_chg1_map.get(
                code,
                float("nan"),
            )
        )

        # --------------------------------------------------
        # A : high-zone stall
        # CHG20 >= 25
        # AND CHG1 < 8
        # AND RSI >= 75
        # AND VolumeRatio <= 2.5
        # --------------------------------------------------

        alert_a = (
            pd.notna(chg20)
            and pd.notna(chg1)
            and pd.notna(rsi)
            and pd.notna(vol)
            and chg20 >= 25
            and chg1 < 8
            and rsi >= 75
            and vol <= 2.5
        )

        # --------------------------------------------------
        # C : one-day spike
        # CHG1 >= 12
        # AND CHG5 < 15
        # AND RSI < 60
        # AND VolumeRatio >= 4
        # --------------------------------------------------

        alert_c = (
            pd.notna(chg1)
            and pd.notna(chg5)
            and pd.notna(rsi)
            and pd.notna(vol)
            and chg1 >= 12
            and chg5 < 15
            and rsi < 60
            and vol >= 4
        )

        # --------------------------------------------------
        # D : extreme overheating
        # (RSI >= 95 AND CHG5 >= 40)
        # OR MA25Deviation >= 80
        # --------------------------------------------------

        alert_d = (
            (
                pd.notna(rsi)
                and pd.notna(chg5)
                and rsi >= 95
                and chg5 >= 40
            )
            or
            (
                pd.notna(ma25dev)
                and ma25dev >= 80
            )
        )

        # --------------------------------------------------
        # F : previous-day surge -> deceleration
        # PREV_CHG1 >= 10
        # AND CHG1 < 8
        # --------------------------------------------------

        alert_f = (
            pd.notna(prev_chg1)
            and pd.notna(chg1)
            and prev_chg1 >= 10
            and chg1 < 8
        )

        # --------------------------------------------------
        # H2 : low score + weak 20-day volume
        # SCORE <= 2
        # AND VolumeRatio20 < 3
        # --------------------------------------------------

        alert_h2 = (
            pd.notna(score)
            and pd.notna(vol20)
            and score <= 2
            and vol20 < 3
        )

        comments = []

        if alert_a:
            comments.append(
                "\u9ad8\u5024\u570f\u5931\u901f"
            )

        if alert_c:
            comments.append(
                "\u5358\u65e5\u5439\u304d\u4e0a\u304c\u308a"
            )

        if alert_d:
            comments.append(
                "\u6975\u7aef\u904e\u71b1"
            )

        if alert_f:
            comments.append(
                "\u524d\u65e5\u6025\u9a30\u304b\u3089\u306e\u5931\u901f"
            )

        if alert_h2:
            comments.append(
                "\u4f4e\u30b9\u30b3\u30a2\u30fb\u4f4e\u51fa\u6765\u9ad8"
            )

        # --------------------------------------------------
        # P5条件
        #
        # 初動スコア 3～4
        # AND 5日騰落率 > 0
        # AND VolumeRatio20 > 1
        # AND A / C / D / F なし
        #
        # H2はP5条件の除外には使わない
        # --------------------------------------------------

        danger4 = (
            alert_a
            or alert_c
            or alert_d
            or alert_f
        )

        p5_condition = (
            pd.notna(score)
            and pd.notna(chg5)
            and pd.notna(vol20)
            and 3 <= score <= 4
            and chg5 > 0
            and vol20 > 1
            and not danger4
        )

        result[code] = {
            "avoid": (
                "\u56de\u907f"
                if comments
                else "\uff0d"
            ),
            "comment": (
                " / ".join(comments)
            ),
            "p5_condition": p5_condition,
        }

    return result



# ==========================================================
# TOP20 Yahoo???? ??
#
# ??????????
# ?TOP20????
# ??????CSV?Yahoo????????
# ??? / ??????Yahoo???
# ==========================================================

YAHOO_CREDIT_DIR = Path(
    "data/yahoo_credit"
)


def _expected_yahoo_credit_date():
    """
    Yahoo??????????????

    ?????????????
    ?????????????????????

    ?????18?????
    1????????????????

    ?:
        ??18??? -> ?????
        ??         -> ???1??????

    Yahoo?????????????????
    ?????????????
    """

    now = pd.Timestamp.now()

    today = now.normalize()

    days_since_friday = (
        today.weekday() - 4
    ) % 7

    candidate = (
        today
        - pd.Timedelta(
            days=days_since_friday
        )
    )

    # candidate(??)?????
    # ????????? = ???18?
    publish_time = (
        candidate
        + pd.Timedelta(days=4)
        + pd.Timedelta(hours=18)
    )

    if now < publish_time:

        candidate = (
            candidate
            - pd.Timedelta(days=7)
        )

    return candidate.normalize()


def _get_credit_csv_latest_date(
    code,
):
    """
    data/yahoo_credit/{code}.csv ?
    ?????????

    ?????????????????
    ?????????????2?????
    """

    path = (
        YAHOO_CREDIT_DIR
        / f"{code}.csv"
    )

    if not path.exists():
        return None

    try:

        df = pd.read_csv(
            path,
            encoding="utf-8-sig",
            dtype=str,
        )

    except Exception:

        return None

    if df.empty:
        return None

    # ----------------------------------------------
    # ??????
    # ----------------------------------------------

    date_col = None

    for col in df.columns:

        name = str(col)

        if (
            "日付" in name
            or "date" in name.lower()
        ):

            date_col = col
            break

    # Yahoo??CSV???
    # 2?????
    if date_col is None:

        if len(df.columns) >= 2:

            date_col = df.columns[1]

        else:

            return None


    dates = pd.to_datetime(
        df[date_col],
        errors="coerce",
    ).dropna()

    if dates.empty:
        return None

    return dates.max().normalize()

def _refresh_top20_credit(
    codes,
):
    """
    TOP20銘柄の信用情報を確認する。
    既存CSVの最新日を確認し、
    必要な銘柄だけYahooから更新する。
    """

    if not codes:
        return

    expected_date = (
        _expected_yahoo_credit_date()
    )

    refresh_codes = []

    print()
    print(
        "TOP20 Yahoo信用データ確認..."
    )

    print(
        "信用データ基準日 :",
        expected_date.date()
    )

    for code in codes:

        code = str(
            code
        ).strip()

        latest_date = (
            _get_credit_csv_latest_date(
                code
            )
        )

        if latest_date is None:

            refresh_codes.append(
                code
            )

            print(
                f"{code} : "
                "既存CSVなし -> Yahoo更新"
            )

            continue

        if latest_date < expected_date:

            refresh_codes.append(
                code
            )

            print(
                f"{code} : "
                f"最新={latest_date.date()} "
                "-> Yahoo更新"
            )

            continue

    print()

    print(
        "TOP20銘柄数      :",
        len(codes)
    )

    print(
        "Yahoo更新対象    :",
        len(refresh_codes)
    )

    print(
        "既存CSV利用      :",
        len(codes)
        - len(refresh_codes)
    )

    # ------------------------------------------------
    # 更新が必要な銘柄だけYahooから取得
    # ------------------------------------------------

    if not refresh_codes:

        print(
            "TOP20信用データはすべて最新です。"
        )

        return

    print()
    print(
        "更新対象のYahoo信用データを取得します..."
    )

    # TOP20だけを対象にする。
    # download_credit_batch() を使用し、
    # Yahooの500エラー対策は既存処理に任せる。
    try:

        download_credit_batch(
            refresh_codes
        )

    except KeyboardInterrupt:

        print()
        print(
            "TOP20信用データ更新を中断しました。"
        )

    except Exception as e:

        print(
            f"TOP20信用データ更新エラー: {e}"
        )

def create_top20_sheet(
    workbook,
    initial_move_top20,
    chart_files,
    reason_data_all,
    news_data,
    sheet_name="TOP20",
):
    """
    7点方式で選出された初動スコアTOP20を
    コンパクトに1枚へまとめる。

    表示項目：
        コード
        銘柄名
        終値
        初動スコア
        信用倍率
        売り残増加
        高騰理由
        ニュース
        ニュース日時
        日足チャート

    ※信用情報はTOP20選出には使用しない。
      あくまでTOP20シートへの表示用。
    """

    # ======================================================
    # TOP20の列
    # ======================================================

    columns = [
        "コード",
        "銘柄名",
        "終値",
        "初動スコア",
        "信用倍率",
        "売り残増加",
        "高騰理由",
        "ニュース",
        "ニュース日時",
        "日足チャート",
    ]

    rows = []

    # ======================================================
    # TOP20銘柄の信用情報を取得
    #
    # ここではTOP20選出後に取得する。
    # 初動スコア計算・順位には一切使用しない。
    # ======================================================

    top20_codes = []

    if not initial_move_top20.empty:

        for _, row in initial_move_top20.iterrows():

            code = str(
                get_value(
                    row,
                    "コード",
                    ""
                )
            ).replace(
                ".0",
                ""
            ).strip()

            if code:
                top20_codes.append(code)

    credit_map = {}

    if top20_codes:

        try:

            # ==============================================
            # TOP20銘柄の信用情報を更新
            # ==============================================

            _refresh_top20_credit(
                top20_codes
            )

            # 更新済みCSVから最新信用データを読み込む
            credit_map = load_latest_credit_data(
                codes=top20_codes
            )

        except Exception as e:

            print(
                f"TOP20信用情報取得エラー: {e}"
            )

            credit_map = {}

    # ======================================================
    # データ作成
    # ======================================================

    if not initial_move_top20.empty:

        for _, row in initial_move_top20.iterrows():

            # ------------------------------------------------
            # コード
            # ------------------------------------------------

            code = str(
                get_value(
                    row,
                    "コード",
                    ""
                )
            ).replace(
                ".0",
                ""
            ).strip()

            # ------------------------------------------------
            # 信用情報
            # ------------------------------------------------

            credit = credit_map.get(
                code
            )

            credit_ratio = ""

            sell_increase = ""

            if credit is not None:

                # --------------------------------------------
                # 信用倍率
                # --------------------------------------------

                credit_ratio = get_value(
                    credit,
                    "信用倍率",
                    ""
                )

                # --------------------------------------------
                # 売り残増加
                #
                # 売り残前週比がプラスなら〇
                # --------------------------------------------

                sell_change = get_value(
                    credit,
                    "売り残前週比",
                    ""
                )

                try:

                    if (
                        sell_change != ""
                        and pd.notna(sell_change)
                        and float(sell_change) > 0
                    ):

                        sell_increase = "〇"

                except (
                    TypeError,
                    ValueError,
                ):

                    sell_increase = ""

            # ------------------------------------------------
            # ニュース情報
            # ------------------------------------------------

            reason = reason_data_all.get(
                code,
                {}
            )

            news_list = news_data.get(
                code,
                []
            )

            # ------------------------------------------------
            # 高騰理由
            # ------------------------------------------------

            main_reason = get_reason_value(
                reason,
                [
                    "main_reason",
                    "reason",
                    "summary",
                    "main_material",
                    "material",
                    "headline",
                ],
                "",
            )

            # ------------------------------------------------
            # ニュースタイトル
            # ------------------------------------------------

            news_title = get_reason_value(
                reason,
                [
                    "main_title",
                    "news_title",
                    "title",
                ],
                "",
            )

            if not news_title:

                news_title = get_first_news_title(
                    news_list
                )

            # ------------------------------------------------
            # ニュース日時
            # ------------------------------------------------

            news_date = get_reason_value(
                reason,
                [
                    "published",
                    "date",
                    "news_date",
                ],
                "",
            )

            if not news_date:

                news_date = get_first_news_date(
                    news_list
                )

            # ------------------------------------------------
            # ニュースURL
            # ------------------------------------------------

            news_link = get_reason_value(
                reason,
                [
                    "main_link",
                    "news_url",
                    "url",
                    "link",
                ],
                "",
            )

            if not news_link:

                news_link = get_first_news_url(
                    news_list
                )

            # ------------------------------------------------
            # ニュース表示
            # ------------------------------------------------

            if news_link:

                news_display = (
                    news_title
                    if news_title
                    else "ニュースを見る"
                )

            else:

                news_display = news_title

            # ------------------------------------------------
            # 高騰理由のフォールバック
            # ------------------------------------------------

            if not main_reason:

                reasons = []

                if get_value(
                    row,
                    "InitialMoveSignal",
                    False
                ) is True:

                    reasons.append(
                        "初動シグナル"
                    )

                if get_value(
                    row,
                    "BreakoutSignal",
                    False
                ) is True:

                    reasons.append(
                        "ブレイクアウト"
                    )

                if get_value(
                    row,
                    "New30High",
                    False
                ) is True:

                    reasons.append(
                        "30日高値更新"
                    )

                if get_value(
                    row,
                    "MACD_GC",
                    False
                ) is True:

                    reasons.append(
                        "MACD GC"
                    )

                volume_ratio = get_value(
                    row,
                    "VolumeRatio",
                    ""
                )

                try:

                    volume_ratio_float = float(
                        volume_ratio
                    )

                    if volume_ratio_float >= 1.5:

                        reasons.append(
                            f"出来高{volume_ratio_float:.1f}倍"
                        )

                except Exception:

                    pass

                price_change = get_value(
                    row,
                    "前日比",
                    ""
                )

                try:

                    price_change_float = float(
                        price_change
                    )

                    if price_change_float > 0:

                        reasons.append(
                            f"当日+{price_change_float:.1f}%"
                        )

                except Exception:

                    pass

                if reasons:

                    main_reason = " / ".join(
                        reasons
                    )

                else:

                    main_reason = (
                        "初動スコアによる抽出"
                    )

            # ------------------------------------------------
            # 1行追加
            # ------------------------------------------------

            rows.append(
                {
                    "コード": code,

                    "銘柄名": get_value(
                        row,
                        "銘柄名",
                        ""
                    ),

                    "終値": get_value(
                        row,
                        "終値",
                        ""
                    ),

                    "初動スコア": get_value(
                        row,
                        "初動スコア",
                        ""
                    ),

                    "信用倍率": credit_ratio,

                    "売り残増加": sell_increase,

                    "高騰理由": main_reason,

                    "ニュース": news_display,

                    "ニュース日時": news_date,

                    "日足チャート": "",
                }
            )

    # ======================================================
    # DataFrame作成
    # ======================================================

    output = pd.DataFrame(
        rows,
        columns=columns
    )

    # ======================================================
    # Buy avoidance alert
    #
    # Insert immediately after Initial Score.
    # Initial Score ranking/calculation is unchanged.
    # ======================================================

    alert_map = (
        _build_buy_avoidance_map(
            initial_move_top20
        )
    )

    p5_values = []
    avoid_values = []
    comment_values = []

    if not output.empty:

        for code_value in output[
            "\u30b3\u30fc\u30c9"
        ]:

            code = str(
                code_value
            ).replace(
                ".0",
                ""
            ).strip()

            info = alert_map.get(
                code,
                {
                    "avoid": "\uff0d",
                    "comment": "",
                    "p5_condition": False,
                }
            )

            p5_values.append(
                "\u2605"
                if info.get(
                    "p5_condition",
                    False,
                )
                else ""
            )

            avoid_values.append(
                info["avoid"]
            )

            comment_values.append(
                info["comment"]
            )

        score_position = (
            output.columns.get_loc(
                "\u521d\u52d5\u30b9\u30b3\u30a2"
            )
            + 1
        )

        output.insert(
            score_position,
            "P5",
            p5_values,
        )

        output.insert(
            score_position + 1,
            "\u8cb7\u3044\u56de\u907f",
            avoid_values,
        )

        output.insert(
            score_position + 2,
            "\u56de\u907f\u30b3\u30e1\u30f3\u30c8",
            comment_values,
        )        
        
    # ======================================================
    # シート作成
    # ======================================================

    sheet = dataframe_to_sheet(
        workbook,
        sheet_name,
        output,
    )

    # ======================================================
    # 列幅
    # ======================================================

    set_column_widths(
        sheet,
        {
            "A": 6,     # code
            "B": 14,    # name
            "C": 8,     # close
            "D": 7,     # initial score
            "E": 6,     # P5　condition
            "F": 6,     # buy avoidance
            "G": 12,    # avoidance comment
            "H": 7,     # credit ratio
            "I": 7,     # short increase
            "J": 13,    # reason
            "K": 20,    # news
            "L": 11,    # news date
            "M": 30,    # chart
        }
    )

    # ======================================================
    # 固定
    # ======================================================

    sheet.freeze_panes = "A2"

    # ======================================================
    # オートフィルター
    # ======================================================

    if sheet.max_row >= 1:

        sheet.auto_filter.ref = (
            sheet.dimensions
        )

    # ======================================================
    # 行高
    # ======================================================

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        sheet.row_dimensions[
            row_number
        ].height = 120

    # ======================================================
    # 初動スコア色付け
    # ======================================================

    apply_score_color(
        sheet,
        "初動スコア"
    )

    # ======================================================
    # ヘッダー位置取得
    # ======================================================

    headers = {}

    for cell in sheet[1]:

        headers[
            cell.value
        ] = cell.column

    # ======================================================
    # 信用倍率の表示形式
    #
    # 1未満：
    #   ・太字
    #   ・赤字
    #   ・フォントサイズ拡大
    # ======================================================

    # ======================================================
    # Buy avoidance display
    #
    # "avoid" = red fill + white bold font
    # ======================================================

    if "\u8cb7\u3044\u56de\u907f" in headers:

        avoid_col = headers[
            "\u8cb7\u3044\u56de\u907f"
        ]

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC00000",
            end_color="FFC00000",
        )

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            cell = sheet.cell(
                row=row_number,
                column=avoid_col,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

            if str(
                cell.value
            ).strip() == "\u56de\u907f":

                cell.fill = red_fill

                cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.sz or 11,
                    bold=True,
                    color="FFFFFFFF",
                )

    if "\u56de\u907f\u30b3\u30e1\u30f3\u30c8" in headers:

        comment_col = headers[
            "\u56de\u907f\u30b3\u30e1\u30f3\u30c8"
        ]

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            cell = sheet.cell(
                row=row_number,
                column=comment_col,
            )

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    if "信用倍率" in headers:

        credit_col = headers[
            "信用倍率"
        ]

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            cell = sheet.cell(
                row=row_number,
                column=credit_col
            )

            try:

                if (
                    cell.value != ""
                    and cell.value is not None
                    and pd.notna(cell.value)
                    and float(cell.value) < 1
                ):

                    cell.font = Font(
                        name=cell.font.name,
                        size=14,
                        bold=True,
                        italic=cell.font.italic,
                        color="FF0000",
                    )

            except (
                TypeError,
                ValueError,
            ):

                pass

    
    # ======================================================
    # ニュースリンク
    #
    # 「ニュース」セルそのものをクリック可能にする。
    # 上揃え＋折り返し表示。
    # ======================================================

    if "ニュース" in headers:

        news_col = headers[
            "ニュース"
        ]

        for row_number, (_, row) in enumerate(
            initial_move_top20.iterrows(),
            start=2
        ):

            code = str(
                get_value(
                    row,
                    "コード",
                    ""
                )
            ).replace(
                ".0",
                ""
            ).strip()

            reason = reason_data_all.get(
                code,
                {}
            )

            news_list = news_data.get(
                code,
                []
            )

            news_link = get_reason_value(
                reason,
                [
                    "main_link",
                    "news_url",
                    "url",
                    "link",
                ],
                "",
            )

            if not news_link:

                news_link = get_first_news_url(
                    news_list
                )

            cell = sheet.cell(
                row=row_number,
                column=news_col
            )

            # ----------------------------------------------
            # ニュース列：左揃え・上揃え・折り返し
            # ----------------------------------------------

            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

            # ----------------------------------------------
            # ニュースリンク
            # ----------------------------------------------

            if news_link:

                cell.hyperlink = news_link

                cell.style = (
                    "Hyperlink"
                )

                # Hyperlinkスタイル適用後も
                # 上揃え＋折り返しを維持
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=True,
                )


    # ======================================================
    # ニュース日時
    # ======================================================

    if "ニュース日時" in headers:

        news_date_col = headers[
            "ニュース日時"
        ]

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            cell = sheet.cell(
                row=row_number,
                column=news_date_col
            )

            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )

    # ======================================================
    # 日足チャート貼り付け
    # ======================================================

    if "日足チャート" in headers:

        chart_col = headers[
            "日足チャート"
        ]

        for row_number, (_, row) in enumerate(
            initial_move_top20.iterrows(),
            start=2
        ):

            code = str(
                get_value(
                    row,
                    "コード",
                    ""
                )
            ).replace(
                ".0",
                ""
            ).strip()

            chart_path = chart_files.get(
                code
            )

            if chart_path:

                try:

                    image = Image(
                        str(chart_path)
                    )

                    image.width = 350
                    image.height = 160

                    cell = sheet.cell(
                        row=row_number,
                        column=chart_col
                    )

                    image.anchor = (
                        cell.coordinate
                    )

                    sheet.add_image(
                        image
                    )

                except Exception as e:

                    print(
                        f"[{code}] "
                        f"チャート貼り付けエラー: {e}"
                    )

    # ======================================================
    # チャート列の中央配置
    # ======================================================

    if chart_col is not None:

        for row_number in range(
            2,
            sheet.max_row + 1
        ):

            cell = sheet.cell(
                row_number,
                chart_col
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    # ======================================================
    # 完了
    # ======================================================

    return sheet

# ==========================================================
# ニュースシート
# ==========================================================

def create_news_sheet(
    workbook,
    initial_move_top20,
    news_data,
):
    """
    初動TOP20の詳細ニュース。

    TOP20本体とは別の補助シート。
    """

    rows = []

    if not initial_move_top20.empty:

        for _, row in initial_move_top20.iterrows():

            code = str(
                get_value(
                    row,
                    "コード",
                    ""
                )
            ).replace(".0", "")
            name = get_value(
                row,
                "銘柄名",
                ""
            )

            score = get_value(
                row,
                "初動スコア",
                ""
            )

            news_list = news_data.get(
                code,
                []
            )

            if not news_list:

                rows.append(
                    {
                        "コード": code,
                        "銘柄名": name,
                        "初動スコア": score,
                        "ニュースタイトル": "",
                        "ニュース概要": "",
                        "ニュース日時": "",
                        "ニュースURL": "",
                    }
                )

                continue

            for news in news_list:

                if not isinstance(
                    news,
                    dict
                ):
                    continue

                rows.append(
                    {
                        "コード": code,
                        "銘柄名": name,
                        "初動スコア": score,
                        "ニュースタイトル": news.get(
                            "title",
                            ""
                        ),
                        "ニュース概要": news.get(
                            "summary",
                            news.get(
                                "description",
                                ""
                            )
                        ),
                        "ニュース日時": news.get(
                            "published",
                            news.get(
                                "date",
                                ""
                            )
                        ),
                        "ニュースURL": news.get(
                            "url",
                            news.get(
                                "link",
                                ""
                            )
                        ),
                    }
                )

    news_df = pd.DataFrame(
        rows
    )

    sheet = dataframe_to_sheet(
        workbook,
        "TOP20ニュース",
        news_df,
    )

    set_column_widths(
        sheet,
        {
            "A": 7,    # コード
            "B": 15,   # 銘柄名
            "C": 8,    # 終値
            "D": 7,    # 初動スコア
            "E": 22,   # 高騰理由
            "F": 22,   # ニュース
            "G": 11,   # ニュース日時
            "H": 25,   # 日足チャート
        }
    )

    if not news_df.empty:

        headers = {}

        for cell in sheet[1]:
            headers[
                cell.value
            ] = cell.column

        if "ニュースURL" in headers:

            url_col = headers[
                "ニュースURL"
            ]

            for row_number in range(
                2,
                sheet.max_row + 1
            ):

                cell = sheet.cell(
                    row_number,
                    url_col
                )

                if not cell.value:
                    continue

                url = str(
                    cell.value
                )

                if (
                    url.startswith(
                        "http://"
                    )
                    or
                    url.startswith(
                        "https://"
                    )
                ):

                    cell.hyperlink = url
                    cell.style = "Hyperlink"

                    cell.value = (
                        "ニュースを見る"
                    )

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        sheet.row_dimensions[
            row_number
        ].height = 105

    return sheet


# ==========================================================
# ニュース理由分析
# ==========================================================

def analyze_top20_news(
    initial_move_top20,
    news_data,
):
    """
    初動TOP20のニュースから
    高騰理由を分析する。
    """

    reason_data_all = {}

    if initial_move_top20.empty:
        return reason_data_all

    for _, row in initial_move_top20.iterrows():

        code = str(
            get_value(
                row,
                "コード",
                ""
            )
        ).replace(".0", "")

        news_list = news_data.get(
            code,
            []
        )

        try:

            reason = analyze_news_reason(
                news_list
            )

        except Exception as e:

            print(
                f"ニュース理由分析ERROR {code} : {e}"
            )

            reason = {}

        if not isinstance(
            reason,
            dict
        ):

            reason = {}

        reason_data_all[
            code
        ] = reason

    return reason_data_all


# ==========================================================
# TOP20 CSV
# ==========================================================

def save_top20_csv(
    initial_move_top20,
    top20_csv,
):
    """
    初動スコアTOP20 CSV保存。
    """

    initial_move_top20.to_csv(
        top20_csv,
        index=False,
        encoding="utf-8-sig",
    )

# ==========================================================
# P5早期候補 Excel
# ==========================================================

def create_p5_candidates_sheet(
    workbook,
    p5_candidates,
    sheet_name="P5早期候補",
):
    """
    正式P5条件を満たした早期候補を一覧表示する。

    P5の判定自体は make_p5_candidates() で完了しているため、
    この関数では表示だけを行う。
    """

    columns = [
        "コード",
        "銘柄名",
        "終値",
        "初動スコア",
        "前日比",
        "5日騰落率",
        "20日騰落率",
        "RSI",
        "VolumeRatio",
        "VolumeRatio20",
        "MA25Deviation",
    ]

    if p5_candidates is None:
        output = pd.DataFrame(
            columns=columns
        )

    elif p5_candidates.empty:
        output = pd.DataFrame(
            columns=columns
        )

    else:
        available_columns = [
            column
            for column in columns
            if column in p5_candidates.columns
        ]

        output = p5_candidates[
            available_columns
        ].copy()

    sheet = dataframe_to_sheet(
        workbook,
        sheet_name,
        output,
    )

    sheet.freeze_panes = "C2"

    # 初動スコア色付け
    apply_score_color(
        sheet,
        "初動スコア",
    )

    # 列幅
    widths = {
        "A": 6,    # コード
        "B": 14,   # 銘柄名
        "C": 8,    # 終値
        "D": 7,    # 初動スコア
        "E": 8,    # 前日比
        "F": 8,    # 5日騰落率
        "G": 8,    # 20日騰落率
        "H": 8,    # RSI
        "I": 8,    # VolumeRatio
        "J": 8,    # VolumeRatio20
        "K": 8,    # MA25Deviation
    }

    for column, width in widths.items():
        sheet.column_dimensions[
            column
        ].width = width

    return sheet

# ==========================================================
# P5追跡 Excel
# ==========================================================

def create_p5_tracking_sheet(
    workbook,
    p5_tracking_df,
    sheet_name="P5追跡",
):
    """
    P5候補のDay1 / Day2と最終買い判定を一覧表示する。
    """

    columns = [
        "DetectionDate",
        "Code",
        "Name",
        "BasePrice",
        "InitialScore",
        "Change5",
        "VolumeRatio20",
        "Day1Price",
        "Day1",
        "Day2Price",
        "Day2",
        "Drop",
        "BuyDecision",
        "BuyReason",
    ]

    if (
        p5_tracking_df is None
        or p5_tracking_df.empty
    ):
        output = pd.DataFrame(
            columns=columns
        )

    else:
        output = p5_tracking_df.copy()

        for column in columns:
            if column not in output.columns:
                output[column] = pd.NA

        output = output[
            columns
        ].copy()

        # 新しい検出日を上へ
        output["DetectionDate"] = (
            pd.to_datetime(
                output["DetectionDate"],
                errors="coerce",
            )
        )

        output = (
            output
            .sort_values(
                [
                    "DetectionDate",
                    "InitialScore",
                ],
                ascending=[
                    False,
                    False,
                ],
                na_position="last",
            )
            .reset_index(
                drop=True
            )
        )

    sheet = dataframe_to_sheet(
        workbook,
        sheet_name,
        output,
    )

    sheet.freeze_panes = "D2"

    widths = {
        "A": 13,   # DetectionDate
        "B": 8,    # Code
        "C": 24,   # Name
        "D": 10,   # BasePrice
        "E": 8,    # InitialScore
        "F": 8,    # Change5
        "G": 8,    # VolumeRatio20
        "H": 8,    # Day1Price
        "I": 8,    # Day1
        "J": 8,    # Day2Price
        "K": 8,    # Day2
        "L": 8,    # Drop
        "M": 8,    # BuyDecision
        "N": 20,   # BuyReason
    }

    for column, width in widths.items():
        sheet.column_dimensions[
            column
        ].width = width

    # 日付表示
    for row in range(
        2,
        sheet.max_row + 1,
    ):
        sheet.cell(
            row,
            1,
        ).number_format = "yyyy/mm/dd"

    # 買い / 見送りを見やすくする
    for row in range(
        2,
        sheet.max_row + 1,
    ):

        decision_cell = sheet.cell(
            row,
            13,
        )

        decision = str(
            decision_cell.value
            or ""
        ).strip()

        if decision == "買い":

            decision_cell.font = Font(
                bold=True
            )

        elif decision == "見送り":

            decision_cell.font = Font(
                bold=True
            )

    return sheet

# ==========================================================
# メイン
# ==========================================================

def save_result(df):

    # ======================================================
    # フォルダ
    # ======================================================

    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    CHART_DIR.mkdir(
        parents=True,
        exist_ok=True
    )

    today = datetime.now().strftime(
        "%Y-%m-%d"
    )

    csv_file = (
        RESULT_DIR
        / f"{today}_stock_result.csv"
    )

    excel_file = (
        RESULT_DIR
        / f"{today}_stock_result.xlsx"
    )

    top20_csv = (
        RESULT_DIR
        / f"{today}_top20.csv"
    )

    # ======================================================
    # 全銘柄CSV
    # ======================================================

    df.to_csv(
        csv_file,
        index=False,
        encoding="utf-8-sig",
    )

    # ======================================================
    # 初動スコアTOP20
    # ======================================================

    initial_move_top20 = (
        make_initial_move_top20(
            df
        )
    )

    # ======================================================
    # P5早期候補
    # ======================================================

    p5_candidates = (
        make_p5_candidates(
            df
        )
    )

    print(
        "P5早期候補件数 :",
        len(p5_candidates)
    )

    # ======================================================
    # P5追跡
    # ======================================================

    p5_tracking_df = load_p5_tracking()

    try:

        # 本日の正式P5候補を追跡へ新規登録
        p5_tracking_df = (
            record_p5_candidates(
                p5_candidates
            )
        )

        # 過去に登録したP5候補の
        # Day1 / Day2 / Drop / 買い判定を更新
        p5_tracking_df = (
            update_p5_tracking(
                p5_tracking_df
            )
        )

        print(
            "P5追跡件数 :",
            len(p5_tracking_df)
        )

    except Exception as e:

        print(
            "P5追跡ERROR :",
            e
        )

    # ======================================================
    # 初動追跡
    # ======================================================

    try:

        tracking_df = load_tracking()

        tracking_df = update_tracking_results(
            tracking_df
        )

        record_initial_move(
            initial_move_top20
        )

    except Exception as e:

        print(
            "初動追跡ERROR :",
            e
        )

    # ======================================================
    # ニュース取得
    # ======================================================

    news_data = {}

    if not initial_move_top20.empty:

        try:

            news_data = get_top20_news(
                initial_move_top20
            )

        except Exception as e:

            print(
                "ニュース取得ERROR :",
                e
            )

            news_data = {}

    # ======================================================
    # ニュース理由分析
    # ======================================================

    reason_data_all = (
        analyze_top20_news(
            initial_move_top20,
            news_data,
        )
    )

    # ======================================================
    # チャート
    # ======================================================

    chart_files = {}

    if not initial_move_top20.empty:

        chart_files = create_charts(
            initial_move_top20
        )

    # ======================================================
    # Excel
    # ======================================================

    print()
    print(
        "Excel作成中..."
    )

    # ======================================================
    # Excel読込 / 新規作成
    # ======================================================

    run_time = datetime.now().strftime(
        "%H-%M"
    )

    if excel_file.exists():

        workbook = load_workbook(
            excel_file
        )

        # ----------------------------------------------
        # 旧形式のTOP20が残っていれば
        # Excelファイルの更新時刻を使って保存
        # ----------------------------------------------

        if "TOP20" in workbook.sheetnames:

            file_time = datetime.fromtimestamp(
                excel_file.stat().st_mtime
            ).strftime(
                "%H-%M"
            )

            legacy_name = (
                f"TOP20_{file_time}"
            )

            candidate = legacy_name
            number = 2

            while candidate in workbook.sheetnames:

                candidate = (
                    f"{legacy_name}_{number}"
                )
                number += 1

            workbook["TOP20"].title = (
                candidate
            )

        # ----------------------------------------------
        # 最新結果へ更新するシートだけ削除
        # ----------------------------------------------

        for sheet_name in [
            "全銘柄",
            "初動追跡",
        ]:

            if sheet_name in workbook.sheetnames:

                workbook.remove(
                    workbook[sheet_name]
                )

    else:

        workbook = Workbook()

        default_sheet = (
            workbook.active
        )

        workbook.remove(
            default_sheet
        )

    # ======================================================
    # 今回のTOP20シート名
    # ======================================================

    top20_sheet_name = (
        f"TOP20_{run_time}"
    )

    base_name = top20_sheet_name
    number = 2

    while (
        top20_sheet_name
        in workbook.sheetnames
    ):

        top20_sheet_name = (
            f"{base_name}_{number}"
        )

        number += 1

    # ======================================================
    # TOP20
    # ======================================================

    create_top20_sheet(
        workbook,
        initial_move_top20,
        chart_files,
        reason_data_all,
        news_data,
        sheet_name=top20_sheet_name,
    )

    # ======================================================
    # P5早期候補
    # ======================================================

    if "P5早期候補" in workbook.sheetnames:

        workbook.remove(
            workbook["P5早期候補"]
        )

    create_p5_candidates_sheet(
        workbook,
        p5_candidates,
        sheet_name="P5早期候補",
    )

    # ======================================================
    # P5追跡
    # ======================================================

    if "P5追跡" in workbook.sheetnames:

        workbook.remove(
            workbook["P5追跡"]
        )

    create_p5_tracking_sheet(
        workbook,
        p5_tracking_df,
        sheet_name="P5追跡",
    )

    # ======================================================
    # 全銘柄
    # ======================================================

    # 不要な重複項目を削除
    df = df.drop(
        columns=["信用倍率計算値"],
        errors="ignore",
    )

    # REMOVE_CREDIT_COLUMNS_FROM_ALL_STOCKS
    # Credit information is shown only on TOP20.
    df = df.drop(
        columns=[
            '信用情報日付',
            '売残',
            '買残',
            '売残前週比',
            '買残前週比',
            '信用倍率',
            '信用条件',
            '信用倍率計算値',
        ],
        errors="ignore",
    )

    result_sheet = dataframe_to_sheet(
        workbook,
        "全銘柄",
        df,
    )

    # コード・銘柄名を固定
    result_sheet.freeze_panes = "C2"

    # 初動スコア色付け
    apply_score_color(
        result_sheet,
        "初動スコア",
    )

    # ======================================================
    # 全銘柄シート体裁調整
    # ======================================================

    headers = {}

    for cell in result_sheet[1]:
        headers[cell.value] = cell.column

    # 分析コメント列を広げる
    if "分析コメント" in headers:

        col = get_column_letter(
            headers["分析コメント"]
        )

        result_sheet.column_dimensions[
            col
        ].width = 50

    # 信用情報日付
    if "信用情報日付" in headers:

        col = headers["信用情報日付"]

        # 日付だけ表示
        for row in range(
            2,
            result_sheet.max_row + 1
        ):

            cell = result_sheet.cell(
                row,
                col
            )

            cell.number_format = "yyyy/mm/dd"

            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
            )

        # ######## を防ぐため列幅を確保
        result_sheet.column_dimensions[
            get_column_letter(col)
        ].width = 14

    # ======================================================
    # 初動追跡
    # ======================================================

    tracking_file = Path(
        "data/tracking/initial_move_tracking.csv"
    )

    if tracking_file.exists():

        try:

            tracking_df = pd.read_csv(
                tracking_file,
                encoding="utf-8-sig",
            )

        except Exception:

            tracking_df = pd.DataFrame()

    else:

        tracking_df = pd.DataFrame()

    tracking_sheet = dataframe_to_sheet(
        workbook,
        "初動追跡",
        tracking_df,
    )

    apply_tracking_change_color(
        tracking_sheet
    )

    # ======================================================
    # シート順
    #
    # 最新TOP20
    # P5早期候補
    # P5追跡
    # 過去TOP20（新しい順）
    # 全銘柄
    # 初動追跡
    # ======================================================

    top20_sheet_names = [
        name
        for name in workbook.sheetnames
        if name.startswith("TOP20_")
    ]

    past_top20_sheet_names = sorted(
        [
            name
            for name in top20_sheet_names
            if name != top20_sheet_name
        ],
        reverse=True
    )

    desired_order = (
        [
            top20_sheet_name,
            "P5早期候補",
            "P5追跡",
        ]
        + past_top20_sheet_names
        + [
            "全銘柄",
            "初動追跡",
        ]
    )

    for position, sheet_name in enumerate(
        desired_order
    ):

        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[
            sheet_name
        ]

        workbook._sheets.remove(
            sheet
        )

        workbook._sheets.insert(
            position,
            sheet
        )

    # ======================================================
    # 最新TOP20を最初に表示
    # ======================================================

    if top20_sheet_name in workbook.sheetnames:

        workbook.active = workbook.index(
            workbook[top20_sheet_name]
        )

    # ======================================================
    # Excel保存
    # ======================================================

    try:

        workbook.save(
            excel_file
        )

    except PermissionError:

        print()
        print(
            "=============================================="
        )
        print(
            f"Excelを保存できません: {excel_file}"
        )
        print(
            "Excelファイルが開いている可能性があります。"
        )
        print(
            "Excelを閉じてから再実行してください。"
        )
        print(
            "=============================================="
        )

        return

    # ======================================================
    # TOP20 CSV
    # ======================================================

    save_top20_csv(
        initial_move_top20,
        top20_csv,
    )

    # ======================================================
    # 完了表示
    # ======================================================

    print()
    print(
        "CSV保存   :",
        csv_file
    )

    print(
        "Excel保存 :",
        excel_file
    )

    print(
        "TOP20保存 :",
        top20_csv
    )
